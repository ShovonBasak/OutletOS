import json
from decimal import Decimal

from django.contrib import admin, messages
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse

from catalog.models import Ingredient, Outlet, PackDefinition, Product
from catalog.ai_extraction import (
    LLMUnavailable,
    available as ai_available,
    extract_historic_prep_log,
    extract_historic_stock_in,
)

from .historic_import import import_stock_in_slip, import_prep_log_slip
from .models import (
    DayStartStockCheck,
    DisplayStock,
    OperatingDay,
    PeriodicStockCheck,
    PreparationLog,
    RawStock,
    StockInItem,
    StockInRecord,
)

SESSION_KEY_STOCKIN = "historic_stock_in_extracted"
SESSION_KEY_PREPLOG = "historic_prep_log_extracted"

# Column-detection helper shared between the Excel import view and the API endpoint.
def _excel_find_col(header, *required_groups, exclude=()):
    """Return the first column index where all keywords in any group are present
    and none of the excluded keywords are present.  ``header`` is a list of
    lowercase strings."""
    for required in required_groups:
        if isinstance(required, str):
            required = [required]
        for i, h in enumerate(header):
            if all(k in h for k in required) and not any(k in h for k in exclude):
                return i
    return None


def _excel_parse_rows(rows):
    """Parse openpyxl row data into a list of item dicts and slip-level totals.

    Returns (items, slip_totals) where:
        items       – list of dicts with the standard import fields
        slip_totals – {"subtotal": str, "vat_total": str, "grand_total": str}
    """
    from catalog.models import SupplierProductAlias

    if not rows:
        return [], {}

    header = [
        str(cell).strip().lower() if cell is not None else "" for cell in rows[0]
    ]

    find = _excel_find_col  # alias

    col_name       = find(header, "product", "service", ["name"], exclude=["ingredient", "base"])
    col_qty        = find(header, "quantity", "qty")
    col_unit       = find(header, "unit", exclude=["per unit", "vat", "sd", "rate"])
    col_rate       = find(header, "per unit rate", "unit rate", ["rate"], exclude=["vat", "sd", "total", "per unit"])
    col_total_amt  = find(header, "total amount", exclude=["vat", "tax", "incl"])
    col_sd_rate    = find(header, ["sd", "rate"])
    col_sd_amt     = find(header, ["sd", "amount"])
    col_vat_rate   = find(header, ["vat", "rate"], ["tax", "rate"])
    col_vat_amt    = find(header, ["vat", "amount"], ["tax", "amount"])
    col_line_total = find(header, "total value", "grand total", ["total", "vat"])

    if col_name is None or col_qty is None:
        return None, {}  # caller checks for None → column error

    # Ingredient lookup
    all_ingredients = {
        ing.name.lower(): ing
        for ing in Ingredient.objects.filter(is_active=True)
    }
    all_aliases = {
        alias.alias_text.lower(): alias.ingredient
        for alias in SupplierProductAlias.objects.filter(is_active=True).select_related("ingredient")
    }

    def _match(text):
        lower = text.strip().lower()
        if lower in all_aliases:
            return all_aliases[lower]
        if lower in all_ingredients:
            return all_ingredients[lower]
        try:
            from rapidfuzz import fuzz, process as rfp
            res = rfp.extractOne(lower, list(all_ingredients), scorer=fuzz.token_sort_ratio)
            if res and res[1] >= 80:
                return all_ingredients[res[0]]
        except ImportError:
            pass
        return None

    def _cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def _s(val):
        if val is None:
            return ""
        s = str(val).strip().replace(",", "")
        return "" if s in ("", "None") else s

    _SKIP = frozenset(["grand total", "sub total", "subtotal", "vat total", "tax total"])
    slip_totals = {"subtotal": "", "vat_total": "", "grand_total": ""}
    items = []

    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        name_raw = _cell(row, col_name)
        if name_raw is None or str(name_raw).strip() == "":
            continue
        name_text = str(name_raw).strip()
        lower = name_text.lower()

        # Capture slip-level totals from summary rows then skip
        if "grand total" in lower:
            v = _cell(row, col_line_total) or _cell(row, col_total_amt)
            if v:
                slip_totals["grand_total"] = _s(v)
            continue
        if "sub total" in lower or "subtotal" in lower:
            v = _cell(row, col_total_amt) or _cell(row, col_line_total)
            if v:
                slip_totals["subtotal"] = _s(v)
            continue
        if "vat total" in lower or "tax total" in lower:
            v = _cell(row, col_vat_amt) or _cell(row, col_line_total)
            if v:
                slip_totals["vat_total"] = _s(v)
            continue
        if any(kw in lower for kw in _SKIP):
            continue

        qty_str = _s(_cell(row, col_qty))
        if not qty_str:
            continue

        unit_raw = _cell(row, col_unit)
        unit = "PIECE" if (unit_raw and str(unit_raw).strip().upper() == "PIECE") else "PACK"

        ingredient = _match(name_text)
        ing_id  = ingredient.pk if ingredient else None
        pack    = ingredient.active_pack() if ingredient else None
        pack_id = pack.pk if pack else None

        rate       = _s(_cell(row, col_rate))
        total_amt  = _s(_cell(row, col_total_amt))
        sd_rate    = _s(_cell(row, col_sd_rate))
        sd_amt     = _s(_cell(row, col_sd_amt))
        vat_rate   = _s(_cell(row, col_vat_rate))
        vat_amt    = _s(_cell(row, col_vat_amt))
        line_total = _s(_cell(row, col_line_total))

        unit_price = ""
        try:
            lt, qt = float(line_total), float(qty_str)
            if lt and qt:
                unit_price = str(round(lt / qt, 4))
        except (ValueError, TypeError):
            pass

        items.append({
            "raw_text":               name_text,
            "matched_ingredient_name": ingredient.name if ingredient else "",
            "ingredient_id":           ing_id,
            "pack_definition_id":      pack_id,
            "quantity":                qty_str,
            "unit":                    unit,
            "unit_price":              unit_price,
            "rate":                    rate,
            "total_amount":            total_amt,
            "sd_rate":                 sd_rate,
            "sd_amount":               sd_amt,
            "vat_rate":                vat_rate,
            "vat_amount":              vat_amt,
            "line_total":              line_total,
        })

    return items, slip_totals


class _DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_images(files) -> list[bytes]:
    out = []
    for f in files:
        f.seek(0)
        out.append(f.read())
    return out


def _resolve_ingredient(name: str | None) -> tuple[int | None, int | None]:
    """Return (ingredient_id, pack_definition_id) for the matched name, or (None, None)."""
    if not name:
        return None, None
    try:
        ing = Ingredient.objects.get(name__iexact=name.strip(), is_active=True)
        pack = ing.pack_definitions.filter(effective_to__isnull=True).first()
        return ing.pk, (pack.pk if pack else None)
    except Ingredient.DoesNotExist:
        return None, None


def _resolve_product(name: str | None) -> int | None:
    """Return product_id for the matched name, or None."""
    if not name:
        return None
    try:
        return Product.objects.get(name__iexact=name.strip(), is_active=True).pk
    except Product.DoesNotExist:
        return None


# ---------------------------------------------------------------------------
# StockInRecord admin (+ historic import wizard)
# ---------------------------------------------------------------------------

class StockInItemInline(admin.TabularInline):
    model = StockInItem
    extra = 0


@admin.register(StockInRecord)
class StockInRecordAdmin(admin.ModelAdmin):
    list_display = ["id", "outlet", "stock_in_date", "status", "submitted_by"]
    list_filter = ["status", "outlet"]
    inlines = [StockInItemInline]
    change_list_template = "admin/stock/stockinrecord/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "import-historic/",
                self.admin_site.admin_view(self.import_historic_upload),
                name="stock_stockinrecord_import_historic",
            ),
            path(
                "import-historic/confirm/",
                self.admin_site.admin_view(self.import_historic_confirm),
                name="stock_stockinrecord_import_historic_confirm",
            ),
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel_upload),
                name="stock_stockinrecord_import_excel",
            ),
            path(
                "sample-excel/",
                self.admin_site.admin_view(self.download_sample_excel),
                name="stock_stockinrecord_sample_excel",
            ),
            path(
                "ocr-debug/",
                self.admin_site.admin_view(self.ocr_debug),
                name="stock_stockinrecord_ocr_debug",
            ),
        ]
        return custom + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["import_historic_url"] = reverse(
            "admin:stock_stockinrecord_import_historic"
        )
        extra_context["import_excel_url"] = reverse(
            "admin:stock_stockinrecord_import_excel"
        )
        extra_context["sample_excel_url"] = reverse(
            "admin:stock_stockinrecord_sample_excel"
        )
        extra_context["ocr_debug_url"] = reverse(
            "admin:stock_stockinrecord_ocr_debug"
        )
        return super().changelist_view(request, extra_context)

    # -- Excel import (no OCR; data already structured) -----------------------

    def download_sample_excel(self, request):
        """Serve a blank Excel template matching the CP Bangladesh invoice layout."""
        import io
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock In"

        headers = [
            "Product / Service Name",
            "Quantity",
            "Unit (PACK/PIECE)",
            "Per Unit Rate (BDT)",
            "Total Amount (BDT)",
            "SD Rate (%)",
            "SD Amount (BDT)",
            "VAT Rate (%)",
            "VAT Amount (BDT)",
            "Total Value incl. VAT & TAX (BDT)",
        ]
        fill = PatternFill(start_color="7A2420", end_color="7A2420", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        examples = [
            ["Chicken Thigh (Raw)", 2, "PACK", 850.00, 1700.00, "", "", 15, 255.00, 1955.00],
            ["Mayonnaise (Portion)", 100, "PIECE", 5.00, 500.00, "", "", 15, 75.00, 575.00],
        ]
        for r, row_data in enumerate(examples, 2):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val if val != "" else None)

        for col, width in enumerate([38, 12, 18, 22, 22, 14, 18, 14, 18, 38], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 42

        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="stock_in_template.xlsx"'
        return resp

    def import_excel_upload(self, request):
        """Step 1: upload Excel → parse → store in session → redirect to confirm."""
        import io
        import openpyxl

        outlets = Outlet.objects.filter(is_active=True)
        ctx = {
            **self.admin_site.each_context(request),
            "title": "Import Stock-In from Excel",
            "outlets": outlets,
            "sample_excel_url": reverse("admin:stock_stockinrecord_sample_excel"),
            "opts": StockInRecord._meta,
        }

        if request.method == "GET":
            return render(request, "admin/stock/import_stock_in_excel.html", ctx)

        outlet_id      = request.POST.get("outlet")
        slip_date_str  = request.POST.get("slip_date", "").strip()
        invoice_number = request.POST.get("invoice_number", "").strip()
        excel_file     = request.FILES.get("excel_file")

        if not outlet_id or not slip_date_str or not excel_file:
            messages.error(request, "Outlet, date, and Excel file are all required.")
            return render(request, "admin/stock/import_stock_in_excel.html", ctx)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        except Exception as exc:
            messages.error(request, f"Could not open Excel file: {exc}")
            return render(request, "admin/stock/import_stock_in_excel.html", ctx)

        rows = list(wb.active.iter_rows(values_only=True))
        items, slip_totals = _excel_parse_rows(rows)

        if items is None:
            messages.error(
                request,
                "Could not detect Name and Quantity columns. "
                "Download the template and use its column headers.",
            )
            return render(request, "admin/stock/import_stock_in_excel.html", ctx)

        if not items:
            messages.error(request, "No item rows found in the Excel file.")
            return render(request, "admin/stock/import_stock_in_excel.html", ctx)

        # Store in the same session format the OCR confirm view expects.
        request.session[SESSION_KEY_STOCKIN] = {
            "outlet_id": outlet_id,
            "slips": [{
                "image_name":     excel_file.name,
                "date":           slip_date_str,
                "invoice_number": invoice_number,
                "subtotal":       slip_totals.get("subtotal", ""),
                "vat_total":      slip_totals.get("vat_total", ""),
                "grand_total":    slip_totals.get("grand_total", ""),
                "items":          items,
            }],
        }
        return HttpResponseRedirect(
            reverse("admin:stock_stockinrecord_import_historic_confirm")
        )

    # -- OCR diagnostics ------------------------------------------------------

    def ocr_debug(self, request):
        """Upload one slip image and see the raw OCR output at every pipeline stage."""
        import io as _io
        from stock.ocr import available as paddle_available
        from stock.ocr._exceptions import PaddleOcrUnavailable, PreprocessingError

        ctx = {
            **self.admin_site.each_context(request),
            "title": "OCR Debug",
            "opts": StockInRecord._meta,
            "paddle_available": paddle_available(),
            "debug": None,
        }

        if request.method == "POST" and request.FILES.get("image"):
            img_bytes = request.FILES["image"].read()
            debug = {"filename": request.FILES["image"].name, "stages": []}

            try:
                from stock.ocr.preprocessing import load_and_preprocess
                from stock.ocr.engine import run, run_text_ocr
                from stock.ocr.slip_parser import parse_structure_result, parse_text_ocr_lines

                image_bgr = load_and_preprocess(_io.BytesIO(img_bytes))
                debug["image_shape"] = list(image_bgr.shape)

                # Stage 1: PP-Structure
                try:
                    pp_result = run(image_bgr)
                    regions = []
                    for r in pp_result:
                        regions.append({
                            "type": r.get("type"),
                            "bbox": r.get("bbox"),
                            "res_summary": str(r.get("res", ""))[:200],
                        })
                    pp_parsed = parse_structure_result(pp_result)
                    debug["stages"].append({
                        "name": "PP-Structure",
                        "regions": regions,
                        "parsed": pp_parsed,
                    })
                except PaddleOcrUnavailable as e:
                    debug["stages"].append({"name": "PP-Structure", "error": str(e)})
                    pp_parsed = {"items": [], "date": None}

                # Stage 2: basic text OCR (only if PP-Structure found nothing)
                try:
                    ocr_lines = run_text_ocr(image_bgr)
                    line_summary = [
                        {"text": (l[1][0] if isinstance(l[1], (list, tuple)) else str(l[1])),
                         "bbox_y": round(sum(p[1] for p in l[0]) / 4, 1),
                         "bbox_x": round(sum(p[0] for p in l[0]) / 4, 1)}
                        for l in ocr_lines if l
                    ]
                    text_parsed = parse_text_ocr_lines(ocr_lines)
                    debug["stages"].append({
                        "name": "Text OCR",
                        "lines": line_summary,
                        "parsed": text_parsed,
                    })
                except PaddleOcrUnavailable as e:
                    debug["stages"].append({"name": "Text OCR", "error": str(e)})

            except PreprocessingError as e:
                debug["error"] = f"Preprocessing failed: {e}"

            ctx["debug"] = debug

        return render(request, "admin/stock/ocr_debug.html", ctx)

    # -- Step 1: upload + OCR extract (PaddleOCR only — no Claude) ------------

    def import_historic_upload(self, request):
        use_ai = ai_available()

        outlets = Outlet.objects.filter(is_active=True)
        ctx = {
            **self.admin_site.each_context(request),
            "title": "Import Historic Stock-In Slips",
            "outlets": outlets,
            "ai_available": use_ai,
            "opts": StockInRecord._meta,
        }

        if request.method == "GET":
            return render(request, "admin/stock/import_stock_in.html", ctx)

        # POST: extract from uploaded images
        outlet_id    = request.POST.get("outlet")
        images_files = request.FILES.getlist("images")

        if not outlet_id or not images_files:
            messages.error(request, "Please select an outlet and upload at least one slip image.")
            return render(request, "admin/stock/import_stock_in.html", ctx)

        if not use_ai:
            messages.error(
                request,
                "ANTHROPIC_API_KEY is not set. "
                "Set it in your .env file to enable invoice extraction.",
            )
            return render(request, "admin/stock/import_stock_in.html", ctx)

        # Get known ingredient names for matching
        known_names = list(
            Ingredient.objects.filter(is_active=True).values_list("name", flat=True)
        )

        slips = []
        for f in images_files:
            img_bytes_data = f.read()

            try:
                parsed = extract_historic_stock_in([img_bytes_data], known_names)
            except LLMUnavailable as exc:
                messages.warning(
                    request,
                    f"{f.name}: Extraction failed — {exc}. "
                    "The slip has been added with empty fields; please fill in manually.",
                )
                parsed = {"date": None, "subtotal": None,
                          "vat_total": None, "grand_total": None, "items": []}
            except Exception as exc:
                messages.warning(request, f"{f.name}: Could not read image — {exc}. Skipped.")
                continue

            if not parsed.get("items") and not parsed.get("date"):
                messages.warning(
                    request,
                    f"{f.name}: No structured data could be extracted. "
                    "The slip has been added with empty fields; please fill in manually.",
                )

            items = []
            for it in parsed.get("items", []):
                matched_name = it.get("matched_ingredient") or ""
                ing_id, pack_id = _resolve_ingredient(matched_name)
                qty_val = it.get("quantity")
                line_total_val = it.get("line_total")
                unit_price_val = None
                if qty_val and line_total_val and float(qty_val) > 0:
                    unit_price_val = round(float(line_total_val) / float(qty_val), 4)
                items.append({
                    "raw_text":               it.get("raw_text", ""),
                    "matched_ingredient_name": matched_name,
                    "ingredient_id":           ing_id,
                    "pack_definition_id":      pack_id,
                    "quantity":                str(qty_val or ""),
                    "unit":                    it.get("unit", "PACK"),
                    "unit_price":              str(unit_price_val or ""),
                    "rate":                    str(it.get("rate") or ""),
                    "total_amount":            str(it.get("total_amount") or ""),
                    "sd_rate":                 str(it.get("sd_rate") or ""),
                    "sd_amount":               str(it.get("sd_amount") or ""),
                    "vat_rate":                str(it.get("vat_rate") or ""),
                    "vat_amount":              str(it.get("vat_amount") or ""),
                    "line_total":              str(it.get("line_total") or ""),
                })

            slips.append({
                "image_name":  f.name,
                "date":        parsed.get("date") or "",
                "subtotal":    str(parsed.get("subtotal") or ""),
                "vat_total":   str(parsed.get("vat_total") or ""),
                "grand_total": str(parsed.get("grand_total") or ""),
                "items":       items,
            })

        if not slips:
            messages.error(request, "No data could be extracted from the uploaded images.")
            return render(request, "admin/stock/import_stock_in.html", ctx)

        request.session[SESSION_KEY_STOCKIN] = {
            "outlet_id": outlet_id,
            "slips": slips,
        }

        return HttpResponseRedirect(
            reverse("admin:stock_stockinrecord_import_historic_confirm")
        )

    # -- Step 2: confirm / edit extracted data + import -----------------------

    def import_historic_confirm(self, request):
        all_ingredients = list(
            Ingredient.objects.filter(is_active=True).values("id", "name", "base_unit")
        )
        all_packs = list(
            PackDefinition.objects.filter(effective_to__isnull=True)
            .select_related("ingredient")
            .values("id", "ingredient_id", "pieces_per_pack", "cost_per_pack")
        )

        ctx = {
            **self.admin_site.each_context(request),
            "title": "Confirm & Import Stock-In Slips",
            "opts": StockInRecord._meta,
            "all_ingredients_json": json.dumps(all_ingredients, cls=_DecimalEncoder),
            "all_packs_json": json.dumps(all_packs, cls=_DecimalEncoder),
        }

        if request.method == "GET":
            session_data = request.session.get(SESSION_KEY_STOCKIN)
            if not session_data:
                messages.error(request, "No pending import data found. Please start over.")
                return HttpResponseRedirect(
                    reverse("admin:stock_stockinrecord_import_historic")
                )
            ctx["session_data_json"] = json.dumps(session_data, cls=_DecimalEncoder)
            return render(request, "admin/stock/import_stock_in_confirm.html", ctx)

        # POST: perform the actual import
        session_data = request.session.get(SESSION_KEY_STOCKIN)
        if not session_data:
            messages.error(request, "Session expired. Please start over.")
            return HttpResponseRedirect(
                reverse("admin:stock_stockinrecord_import_historic")
            )

        try:
            outlet = Outlet.objects.get(pk=session_data["outlet_id"])
        except Outlet.DoesNotExist:
            messages.error(request, "Selected outlet no longer exists.")
            return HttpResponseRedirect(
                reverse("admin:stock_stockinrecord_import_historic")
            )

        num_slips = int(request.POST.get("num_slips", 0))
        total_records = 0
        total_warnings = []

        from datetime import date as date_type

        for i in range(num_slips):
            date_str = request.POST.get(f"slip_{i}_date", "").strip()
            if not date_str:
                messages.warning(request, f"Slip {i + 1}: no date provided — skipped.")
                continue

            try:
                slip_date = date_type.fromisoformat(date_str)
            except ValueError:
                messages.warning(request, f"Slip {i + 1}: invalid date '{date_str}' — skipped.")
                continue

            invoice_number = request.POST.get(f"slip_{i}_invoice_number", "").strip()
            slip_totals = {
                "subtotal":    request.POST.get(f"slip_{i}_subtotal") or None,
                "vat_total":   request.POST.get(f"slip_{i}_vat_total") or None,
                "grand_total": request.POST.get(f"slip_{i}_grand_total") or None,
            }

            num_items = int(request.POST.get(f"slip_{i}_num_items", 0))
            items = []
            for j in range(num_items):
                if not request.POST.get(f"item_{i}_{j}_include"):
                    continue
                qty_raw = request.POST.get(f"item_{i}_{j}_quantity", "0")
                try:
                    qty = float(qty_raw)
                except (ValueError, TypeError):
                    qty = 0
                if qty <= 0:
                    continue
                items.append({
                    "ingredient_id":      request.POST.get(f"item_{i}_{j}_ingredient_id") or None,
                    "raw_text":           request.POST.get(f"item_{i}_{j}_raw_text", ""),
                    "quantity":           qty,
                    "unit":               request.POST.get(f"item_{i}_{j}_unit", "PACK"),
                    "pack_definition_id": request.POST.get(f"item_{i}_{j}_pack_definition_id") or None,
                    "unit_price":         request.POST.get(f"item_{i}_{j}_unit_price") or None,
                    "rate":               request.POST.get(f"item_{i}_{j}_rate") or None,
                    "total_amount":       request.POST.get(f"item_{i}_{j}_total_amount") or None,
                    "sd_rate":            request.POST.get(f"item_{i}_{j}_sd_rate") or None,
                    "sd_amount":          request.POST.get(f"item_{i}_{j}_sd_amount") or None,
                    "vat_rate":           request.POST.get(f"item_{i}_{j}_vat_rate") or None,
                    "vat_amount":         request.POST.get(f"item_{i}_{j}_vat_amount") or None,
                    "line_total":         request.POST.get(f"item_{i}_{j}_line_total") or None,
                })

            if not items:
                continue

            record, warnings = import_stock_in_slip(
                outlet, slip_date, items, request.user,
                slip_totals=slip_totals, invoice_number=invoice_number,
            )
            total_records += 1
            total_warnings.extend(warnings)

        del request.session[SESSION_KEY_STOCKIN]

        if total_records:
            messages.success(
                request,
                f"Imported {total_records} stock-in record(s). RawStock has been updated.",
            )
        for w in total_warnings:
            messages.warning(request, w)

        return HttpResponseRedirect(reverse("admin:stock_stockinrecord_changelist"))


# ---------------------------------------------------------------------------
# PreparationLog admin (+ historic import wizard)
# ---------------------------------------------------------------------------

@admin.register(PreparationLog)
class PreparationLogAdmin(admin.ModelAdmin):
    list_display = [
        "product", "outlet", "source", "prep_unit", "packs_used",
        "pieces_prepared", "wastage_pieces", "timestamp",
    ]
    list_filter = ["outlet", "source", "prep_unit"]
    change_list_template = "admin/stock/preparationlog/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "import-historic/",
                self.admin_site.admin_view(self.import_historic_upload),
                name="stock_preparationlog_import_historic",
            ),
            path(
                "import-historic/confirm/",
                self.admin_site.admin_view(self.import_historic_confirm),
                name="stock_preparationlog_import_historic_confirm",
            ),
        ]
        return custom + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["import_historic_url"] = reverse(
            "admin:stock_preparationlog_import_historic"
        )
        return super().changelist_view(request, extra_context)

    # -- Step 1: upload + AI extract ------------------------------------------

    def import_historic_upload(self, request):
        outlets = Outlet.objects.filter(is_active=True)
        ctx = {
            **self.admin_site.each_context(request),
            "title": "Import Historic Prep Log Slips",
            "outlets": outlets,
            "ai_available": ai_available(),
            "opts": PreparationLog._meta,
        }

        if request.method == "GET":
            return render(request, "admin/stock/import_prep_log.html", ctx)

        outlet_id = request.POST.get("outlet")
        images_files = request.FILES.getlist("images")

        if not outlet_id or not images_files:
            messages.error(request, "Please select an outlet and upload at least one slip image.")
            return render(request, "admin/stock/import_prep_log.html", ctx)

        if not ai_available():
            messages.error(request, "ANTHROPIC_API_KEY is not set — AI extraction unavailable.")
            return render(request, "admin/stock/import_prep_log.html", ctx)

        known_products = list(Product.objects.filter(is_active=True).values_list("name", flat=True))

        slips = []
        for f in images_files:
            try:
                img_bytes = [f.read()]
                result = extract_historic_prep_log(img_bytes, known_products)
            except LLMUnavailable as exc:
                messages.warning(request, f"{f.name}: AI extraction failed — {exc}")
                continue

            items = []
            for it in result.get("items", []):
                product_id = _resolve_product(it.get("matched_product"))
                items.append({
                    "raw_text": it.get("raw_text", ""),
                    "matched_product_name": it.get("matched_product") or "",
                    "product_id": product_id,
                    "pieces_prepared": str(it.get("pieces_prepared") or ""),
                    "source": it.get("source", "FRESH"),
                })
            slips.append({
                "image_name": f.name,
                "date": result.get("date") or "",
                "items": items,
            })

        if not slips:
            messages.error(request, "No data could be extracted from the uploaded images.")
            return render(request, "admin/stock/import_prep_log.html", ctx)

        request.session[SESSION_KEY_PREPLOG] = {
            "outlet_id": outlet_id,
            "slips": slips,
        }

        return HttpResponseRedirect(
            reverse("admin:stock_preparationlog_import_historic_confirm")
        )

    # -- Step 2: confirm / edit extracted data + import -----------------------

    def import_historic_confirm(self, request):
        all_products = list(
            Product.objects.filter(is_active=True).values("id", "name", "category")
        )

        ctx = {
            **self.admin_site.each_context(request),
            "title": "Confirm & Import Prep Log Slips",
            "opts": PreparationLog._meta,
            "all_products_json": json.dumps(all_products, cls=_DecimalEncoder),
        }

        if request.method == "GET":
            session_data = request.session.get(SESSION_KEY_PREPLOG)
            if not session_data:
                messages.error(request, "No pending import data found. Please start over.")
                return HttpResponseRedirect(
                    reverse("admin:stock_preparationlog_import_historic")
                )
            ctx["session_data_json"] = json.dumps(session_data, cls=_DecimalEncoder)
            return render(request, "admin/stock/import_prep_log_confirm.html", ctx)

        # POST: perform the actual import
        session_data = request.session.get(SESSION_KEY_PREPLOG)
        if not session_data:
            messages.error(request, "Session expired. Please start over.")
            return HttpResponseRedirect(
                reverse("admin:stock_preparationlog_import_historic")
            )

        try:
            outlet = Outlet.objects.get(pk=session_data["outlet_id"])
        except Outlet.DoesNotExist:
            messages.error(request, "Selected outlet no longer exists.")
            return HttpResponseRedirect(
                reverse("admin:stock_preparationlog_import_historic")
            )

        num_slips = int(request.POST.get("num_slips", 0))
        total_logs = 0
        total_warnings = []

        for i in range(num_slips):
            date_str = request.POST.get(f"slip_{i}_date", "").strip()
            if not date_str:
                messages.warning(request, f"Slip {i + 1}: no date provided — skipped.")
                continue

            from datetime import date as date_type
            try:
                slip_date = date_type.fromisoformat(date_str)
            except ValueError:
                messages.warning(request, f"Slip {i + 1}: invalid date '{date_str}' — skipped.")
                continue

            num_items = int(request.POST.get(f"slip_{i}_num_items", 0))
            items = []
            for j in range(num_items):
                if not request.POST.get(f"item_{i}_{j}_include"):
                    continue
                try:
                    pieces = int(request.POST.get(f"item_{i}_{j}_pieces_prepared", 0))
                except (ValueError, TypeError):
                    pieces = 0
                if pieces <= 0:
                    continue
                items.append({
                    "product_id": request.POST.get(f"item_{i}_{j}_product_id") or None,
                    "raw_text": request.POST.get(f"item_{i}_{j}_raw_text", ""),
                    "pieces_prepared": pieces,
                    "source": request.POST.get(f"item_{i}_{j}_source", "FRESH"),
                })

            if not items:
                continue

            logs, warnings = import_prep_log_slip(outlet, slip_date, items, request.user)
            total_logs += len(logs)
            total_warnings.extend(warnings)

        del request.session[SESSION_KEY_PREPLOG]

        if total_logs:
            messages.success(
                request,
                f"Imported {total_logs} preparation log entry(s). DisplayStock and RawStock updated.",
            )
        for w in total_warnings:
            messages.warning(request, w)

        return HttpResponseRedirect(reverse("admin:stock_preparationlog_changelist"))


# ---------------------------------------------------------------------------
# Other stock models
# ---------------------------------------------------------------------------

class DayStartStockCheckInline(admin.TabularInline):
    model = DayStartStockCheck
    extra = 0


@admin.register(OperatingDay)
class OperatingDayAdmin(admin.ModelAdmin):
    list_display = ["date", "outlet", "status", "started_by"]
    list_filter = ["status", "outlet"]
    inlines = [DayStartStockCheckInline]


@admin.register(PeriodicStockCheck)
class PeriodicStockCheckAdmin(admin.ModelAdmin):
    list_display = [
        "ingredient", "outlet", "counted_qty", "consumed_since_last_check", "checked_at",
    ]
    list_filter = ["outlet", "ingredient"]


admin.site.register(RawStock)
admin.site.register(DisplayStock)
