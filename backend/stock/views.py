from decimal import Decimal

from django.db import models
from django.db.models import Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from accounts.permissions import IsOwnerOrAdmin
from catalog.models import Ingredient, SupplierProductAlias, TrackingMode
from .extraction import ExtractedLine
from .models import (
    DayStartStockCheck,
    DisplayStock,
    FryerOilChange,
    LineSource,
    OperatingDay,
    OperatingDayStatus,
    PeriodicStockCheck,
    PrepSource,
    PrepUnit,
    PreparationLog,
    RawStock,
    StockInItem,
    StockInRecord,
    StockInStatus,
    UnitCaptured,
)
from .serializers import (
    DayStartStockCheckSerializer,
    DisplayStockPrepSerializer,
    DisplayStockSerializer,
    FryerOilChangeSerializer,
    OperatingDaySerializer,
    OperatingDaySlimSerializer,
    PeriodicStockCheckSerializer,
    PrepLogSlimSerializer,
    PreparationLogSerializer,
    RawStockPrepSerializer,
    RawStockSerializer,
    StockInRecordListSerializer,
    StockInRecordSerializer,
)
from .services import (
    carry_forward_candidates,
    consume_for_preparation,
    get_or_create_today,
    restock_from_preparation,
    stock_in_since,
)


class StockInListPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100


class StockInRecordViewSet(viewsets.ModelViewSet):
    queryset = StockInRecord.objects.all()  # required by DRF router for basename detection
    serializer_class = StockInRecordSerializer
    pagination_class = StockInListPagination

    _FULL_QUERYSET = StockInRecord.objects.prefetch_related(
        "items__ingredient", "items__pack_definition"
    ).select_related("outlet", "submitted_by", "paid_from_account")

    def _list_queryset(self):
        return StockInRecord.objects.select_related(
            "submitted_by", "paid_from_account"
        ).annotate(
            item_count=Count("items", distinct=True),
            slip_item_count=Count(
                "items", filter=Q(items__source="SLIP_EXTRACTED"), distinct=True
            ),
            manual_item_count=Count(
                "items", filter=Q(items__source="MANUAL"), distinct=True
            ),
        )

    def _is_slim(self):
        return (
            self.action == "list"
            and self.request.query_params.get("expand") != "full"
        )

    def get_queryset(self):
        qs = self._list_queryset() if self._is_slim() else self._FULL_QUERYSET
        status_param = self.request.query_params.get("status")
        if status_param:
            qs = qs.filter(status=status_param)
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(invoice_number__icontains=search) | Q(id__icontains=search)
            )
        return qs.order_by("-stock_in_date", "-created_at")

    def get_serializer_class(self):
        if self._is_slim():
            return StockInRecordListSerializer
        return StockInRecordSerializer

    def perform_create(self, serializer):
        serializer.save(submitted_by=self.request.user)

    def _guard_editable(self, record):
        if record.status not in (StockInStatus.DRAFT,):
            raise ValidationError("Only DRAFT records can be edited by staff.")

    def update(self, request, *args, **kwargs):
        self._guard_editable(self.get_object())
        return super().update(request, *args, **kwargs)

    def perform_destroy(self, instance):
        if instance.status != StockInStatus.DRAFT:
            raise ValidationError("Only DRAFT records can be deleted.")
        instance.delete()

    @action(
        detail=True,
        methods=["post"],
        url_path="upload-slip",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_slip(self, request, pk=None):
        """Attach/replace the delivery-slip image (multipart field `slip_image`)."""
        record = self.get_object()
        self._guard_editable(record)
        image = request.FILES.get("slip_image")
        if not image:
            raise ValidationError("No slip_image file provided.")
        record.slip_image = image
        record.save(update_fields=["slip_image"])
        return Response(self.get_serializer(record).data)

    @action(detail=True, methods=["post"])
    def extract(self, request, pk=None):
        """OCR the attached slip image with Claude → create/refresh SLIP_EXTRACTED lines.

        Unmatched lines become 'Unrecognized' rows (ingredient=None) for staff
        to resolve. Existing slip-extracted lines are replaced; manual lines kept.
        """
        from catalog import ai_extraction
        from catalog.ai_extraction import LLMUnavailable

        record = self.get_object()
        self._guard_editable(record)
        if not record.slip_image:
            raise ValidationError("Attach a slip image before extracting.")

        known_names = list(
            Ingredient.objects.filter(is_active=True).values_list("name", flat=True)
        )
        with open(record.slip_image.path, "rb") as fh:
            image_data = fh.read()

        try:
            result = ai_extraction.extract_historic_stock_in([image_data], known_names)
        except LLMUnavailable as exc:
            return Response(
                {
                    "detail": "Claude AI is unavailable — please enter lines manually.",
                    "reason": str(exc),
                    "ocr_available": False,
                },
                status=503,
            )

        def _dec(v):
            if v is None:
                return None
            try:
                return Decimal(str(v))
            except Exception:
                return None

        # Update StockInRecord header fields from the extracted slip data.
        record_update_fields = []
        invoice_number = result.get("invoice_number")
        if invoice_number and not record.invoice_number:
            record.invoice_number = str(invoice_number).strip()
            record_update_fields.append("invoice_number")
        slip_date = result.get("date")
        if slip_date:
            from datetime import date as _date
            try:
                parsed_date = _date.fromisoformat(str(slip_date))
                record.stock_in_date = parsed_date
                record_update_fields.append("stock_in_date")
            except ValueError:
                pass
        for field, key in [
            ("slip_subtotal", "subtotal"),
            ("slip_vat_total", "vat_total"),
            ("slip_grand_total", "grand_total"),
        ]:
            val = _dec(result.get(key))
            if val is not None:
                setattr(record, field, val)
                record_update_fields.append(field)
        if record_update_fields:
            record.save(update_fields=record_update_fields)

        lines: list[ExtractedLine] = []
        for item in result.get("items", []):
            matched_name = item.get("matched_ingredient")
            ingredient_id = None
            pack_definition_id = None
            if matched_name:
                ing = Ingredient.objects.filter(name=matched_name, is_active=True).first()
                if ing:
                    ingredient_id = ing.id
                    pack = ing.pack_definitions.filter(effective_to__isnull=True).first()
                    pack_definition_id = pack.id if pack else None
            qty = item.get("quantity")
            lines.append(ExtractedLine(
                raw_text=item.get("raw_text", ""),
                extracted_quantity=float(qty) if qty is not None else None,
                ingredient_id=ingredient_id,
                pack_definition_id=pack_definition_id,
                unit_captured=item.get("unit", "PACK"),
                rate=item.get("rate"),
                total_amount=item.get("total_amount"),
                sd_rate=item.get("sd_rate"),
                sd_amount=item.get("sd_amount"),
                vat_rate=item.get("vat_rate"),
                vat_amount=item.get("vat_amount"),
                line_total=item.get("line_total"),
                unit_price=item.get("unit_price"),
            ))

        record.items.filter(source=LineSource.SLIP_EXTRACTED).delete()
        for line in lines:
            StockInItem.objects.create(
                stock_in_record=record,
                ingredient_id=line.ingredient_id,
                pack_definition_id=line.pack_definition_id,
                raw_extracted_text=line.raw_text,
                source=LineSource.SLIP_EXTRACTED,
                unit_captured=line.unit_captured,
                extracted_quantity=line.extracted_quantity,
                confirmed_quantity=line.extracted_quantity or Decimal("0"),
                rate=_dec(line.rate),
                total_amount=_dec(line.total_amount),
                sd_rate=_dec(line.sd_rate),
                sd_amount=_dec(line.sd_amount),
                vat_rate=_dec(line.vat_rate),
                vat_amount=_dec(line.vat_amount),
                line_total=_dec(line.line_total),
                unit_price=_dec(line.unit_price),
            )
        record = self.get_queryset().get(pk=record.pk)
        data = self.get_serializer(record).data
        data["extracted_count"] = len(lines)
        return Response(data)

    @action(detail=True, methods=["post"], url_path="resolve-line")
    def resolve_line(self, request, pk=None):
        """Resolve an Unrecognized line: assign an ingredient (existing or new)
        and remember the mapping as a SupplierProductAlias for next time.
        Body: {item, ingredient?|new_ingredient_name?, base_unit?, pack_definition?}."""
        record = self.get_object()
        self._guard_editable(record)
        item = record.items.get(pk=request.data["item"])

        ingredient_id = request.data.get("ingredient")
        if ingredient_id:
            ingredient = Ingredient.objects.get(pk=ingredient_id)
        else:
            name = (request.data.get("new_ingredient_name") or "").strip()
            if not name:
                raise ValidationError("Provide ingredient or new_ingredient_name.")
            ingredient, _ = Ingredient.objects.get_or_create(
                name=name,
                defaults={"base_unit": request.data.get("base_unit", "piece")},
            )
        item.ingredient = ingredient
        pack_definition_id = request.data.get("pack_definition")
        if pack_definition_id:
            item.pack_definition_id = pack_definition_id
        else:
            active = ingredient.active_pack()
            item.pack_definition = active
        item.save()

        # Remember the supplier wording so it auto-resolves next time.
        if item.raw_extracted_text:
            SupplierProductAlias.objects.get_or_create(
                ingredient=ingredient, alias_text=item.raw_extracted_text
            )
        record = self.get_queryset().get(pk=record.pk)
        return Response(self.get_serializer(record).data)

    @action(detail=True, methods=["post"], url_path="set-pack-yield")
    def set_pack_yield(self, request, pk=None):
        """Answer the inline 'how many pieces does 1 pack make?' prompt — creates
        the ingredient's first PackDefinition on the spot. Body:
        {item, pieces_per_pack, cost_per_pack}."""
        from catalog.models import PackDefinition

        record = self.get_object()
        self._guard_editable(record)
        item = record.items.get(pk=request.data["item"])
        if not item.ingredient_id:
            raise ValidationError("Resolve the ingredient before setting pack yield.")
        pack = PackDefinition.objects.create(
            ingredient=item.ingredient,
            pieces_per_pack=Decimal(str(request.data["pieces_per_pack"])),
            cost_per_pack=Decimal(str(request.data.get("cost_per_pack", 0))),
            effective_from=timezone.localdate(),
        )
        item.pack_definition = pack
        item.save(update_fields=["pack_definition"])
        record = self.get_queryset().get(pk=record.pk)
        return Response(self.get_serializer(record).data)

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        """Staff moves DRAFT → PENDING. Blocked if any line is unresolved."""
        record = self.get_object()
        if record.status != StockInStatus.DRAFT:
            raise ValidationError("Only DRAFT records can be submitted.")
        unresolved = record.unresolved_lines
        if unresolved:
            raise ValidationError(
                {
                    "detail": "Resolve all lines before submitting "
                    "(unknown ingredient or missing pack yield).",
                    "unresolved_item_ids": [i.id for i in unresolved],
                }
            )
        record.status = StockInStatus.PENDING
        record.save()

        try:
            from accounts.push import send_push_to_owners
            item_count = record.items.count()
            send_push_to_owners(
                title="Stock-in waiting for approval",
                body=f"{record.submitted_by.name} submitted {item_count} item(s) for {record.stock_in_date}",
                url="/owner/stock-in",
            )
        except Exception:
            pass

        return Response(self.get_serializer(record).data)

    @action(detail=True, methods=["post"], permission_classes=[IsOwnerOrAdmin])
    def approve(self, request, pk=None):
        """Owner approves PENDING → APPROVED; RawStock increments (base units).

        Also price-versions PackDefinition for any line that carries price data
        from the slip so that future COGS calculations use the current purchase cost.
        Accepts optional paid_from_account in request body to override the account.
        """
        from .historic_import import _effective_cost_per_pack, update_pack_definition_cost
        from finance.models import AccountTransaction, FinancialAccount
        from decimal import Decimal

        record = self.get_object()
        if record.status != StockInStatus.PENDING:
            raise ValidationError("Only PENDING records can be approved.")

        # Allow owner to override or set the payment account at approval time.
        account_id = request.data.get("paid_from_account")
        if account_id is not None:
            try:
                record.paid_from_account = FinancialAccount.objects.get(pk=account_id) if account_id else None
            except FinancialAccount.DoesNotExist:
                raise ValidationError("Invalid account.")

        for item in record.items.select_related("ingredient", "pack_definition"):
            if not item.ingredient_id:
                continue
            if item.ingredient.tracking_mode in (TrackingMode.ONE_TIME, TrackingMode.PERIODIC_COUNT):
                continue  # ONE_TIME: cost audit only; PERIODIC_COUNT: tracked via periodic checks
            RawStock.adjust(record.outlet, item.ingredient, item.base_unit_quantity())

            if (item.pack_definition_id
                    and item.unit_captured == UnitCaptured.PACK
                    and item.confirmed_quantity):
                cost = _effective_cost_per_pack(
                    item.unit_price, item.line_total, item.confirmed_quantity
                )
                if cost:
                    update_pack_definition_cost(
                        item.pack_definition_id, cost, record.stock_in_date
                    )

        record.status = StockInStatus.APPROVED
        record.reviewed_by = request.user
        record.reviewed_at = timezone.now()
        record.save()

        # Deduct from the payment account if one is set.
        if record.paid_from_account_id:
            total = record.slip_grand_total or Decimal(
                sum(
                    Decimal(item.line_total or "0")
                    for item in record.items.all()
                )
            )
            if total > 0:
                AccountTransaction.objects.create(
                    account=record.paid_from_account,
                    transaction_type="SUPPLIER_ORDER_DEDUCTION",
                    amount=-total,
                    date=record.stock_in_date,
                    source_type="STOCK_IN_RECORD",
                    source_id=record.id,
                    entered_by=request.user,
                    note=f"Stock-in {record.invoice_number or f'#{record.id}'} approved",
                )

        _initialize_direct_stock(record.outlet)

        # Refresh open closing for today: new stock goes to *remains*, not to sold.
        # Only update products whose ingredients are in THIS approval — after
        # stock_count runs, RawStock = remains * qty_per, so _initialize_direct_stock
        # sets DS = remains for unrelated products. Touching those counts would
        # overwrite available_pieces with remains and collapse derived_walkin_sold to 0.
        from closing.models import DailyClosing, DailyClosingStockCount
        from closing.services import recompute_closing as _recompute_closing
        open_closing = (
            DailyClosing.objects
            .filter(outlet=record.outlet, closing_date=record.stock_in_date)
            .exclude(status="LOCKED")
            .first()
        )
        if open_closing:
            approved_ingredient_ids = {
                item.ingredient_id
                for item in record.items.select_related("ingredient")
                if item.ingredient_id
                and item.ingredient.tracking_mode
                not in (TrackingMode.ONE_TIME, TrackingMode.PERIODIC_COUNT)
            }
            closing_updated = False
            for sc in open_closing.stock_counts.select_related("product").filter(
                product__requires_preparation=False
            ):
                recipe_ingredient_ids = set(
                    sc.product.recipes.values_list("ingredient_id", flat=True)
                )
                if not (recipe_ingredient_ids & approved_ingredient_ids):
                    continue  # this product not in the current stock-in — leave it alone
                ds = DisplayStock.objects.filter(
                    outlet=record.outlet, product=sc.product
                ).first()
                if not ds:
                    continue
                # ds.pieces_available = old_remains + new_stock_in_pieces (just set by
                # _initialize_direct_stock). Add that delta to both counters so that
                # derived_walkin_sold = available − wastage − remains − app_sold is unchanged.
                new_pieces = ds.pieces_available - sc.remains_pieces
                if new_pieces > 0:
                    sc.available_pieces += new_pieces
                    sc.remains_pieces = ds.pieces_available
                    sc.save(update_fields=["available_pieces", "remains_pieces"])
                    closing_updated = True
            if closing_updated:
                _recompute_closing(open_closing)

        return Response(self.get_serializer(record).data)

    @action(detail=True, methods=["post"], permission_classes=[IsOwnerOrAdmin])
    def reject(self, request, pk=None):
        record = self.get_object()
        if record.status != StockInStatus.PENDING:
            raise ValidationError("Only PENDING records can be rejected.")
        record.status = StockInStatus.REJECTED
        record.reviewed_by = request.user
        record.reviewed_at = timezone.now()
        record.notes = request.data.get("notes", record.notes)
        record.save()
        return Response(self.get_serializer(record).data)

    # ── Excel import ─────────────────────────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="sample-excel")
    def sample_excel(self, request):
        """Download a blank Excel template matching the CP Bangladesh invoice layout."""
        import io
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock In"

        headers = [
            "Product Code",
            "Product / Service Name",
            "Quantity",
            "Unit (PACK/PIECE)",
            "MRP per Piece (BDT)",
            "Per Unit Rate (BDT)",
            "Total Amount (BDT)",
            "Discount (BDT)",
            "SD Rate (%)",
            "SD Amount (BDT)",
            "VAT Rate (%)",
            "VAT Amount (BDT)",
            "Total Value incl. Discount, VAT & TAX (BDT)",
        ]
        header_fill = PatternFill(start_color="7A2420", end_color="7A2420", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

        # Example rows: one food slip line, one beverage challan line
        examples = [
            ["",       "Chicken Thigh (Raw)",      2,   "PACK",  "",  850.00, 1700.00, "",    "", "", 15, 255.00, 1955.00],
            ["738803", "CCZS PET 250ML 1X24",       1,  "PACK",  20,  432.00,  432.00, 38.90, "", "", "",     "", 393.10],
        ]
        for r, row_data in enumerate(examples, 2):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val if val != "" else None)

        for col, width in enumerate(
            [16, 38, 12, 18, 18, 22, 22, 16, 14, 18, 14, 18, 42], 1
        ):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 42

        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="stock_in_template.xlsx"'
        return resp

    @action(
        detail=True,
        methods=["post"],
        url_path="import-excel",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_excel(self, request, pk=None):
        """Parse an Excel file and populate StockInItems on a DRAFT record.

        Multipart field: ``excel_file``

        Columns are auto-detected by header text (case-insensitive).  The
        sample template column order is the canonical reference, but any order
        works.  Existing SLIP_EXTRACTED lines are replaced; MANUAL lines kept.

        Returns the updated record plus ``imported_count`` and
        ``unresolved_names`` (ingredient could not be matched — staff must
        resolve before submitting).
        """
        import io
        import openpyxl

        record = self.get_object()
        self._guard_editable(record)

        # Optional metadata fields that update the record before importing lines.
        invoice_number = request.data.get("invoice_number", "").strip()
        stock_in_date  = request.data.get("stock_in_date", "").strip()
        slip_image     = request.FILES.get("slip_image")
        update_fields = []
        if invoice_number:
            record.invoice_number = invoice_number
            update_fields.append("invoice_number")
        if stock_in_date:
            from datetime import date
            try:
                record.stock_in_date = date.fromisoformat(stock_in_date)
                update_fields.append("stock_in_date")
            except ValueError:
                raise ValidationError("Invalid stock_in_date — use YYYY-MM-DD.")
        if slip_image:
            record.slip_image = slip_image
            update_fields.append("slip_image")
        if update_fields:
            record.save(update_fields=update_fields)

        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            raise ValidationError("No excel_file provided.")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
        except Exception as exc:
            raise ValidationError(f"Could not read Excel file: {exc}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=False))  # Cell objects — needed for number_format
        if not rows:
            raise ValidationError("Excel file is empty.")

        header = [
            str(cell.value).strip().lower() if cell.value is not None else "" for cell in rows[0]
        ]

        def _find(required, exclude=()):
            """First column index where all ``required`` substrings are present
            and none of ``exclude`` substrings are present."""
            for i, h in enumerate(header):
                if all(k in h for k in required) and not any(k in h for k in exclude):
                    return i
            return None

        col_sku       = _find(["product code"]) or _find(["sku"]) or _find(["পণ্য কোড"])
        col_name      = _find(["product / service"]) or _find(["product"], ["code", "sku"]) or _find(["service"]) or _find(["name"], ["ingredient", "base"])
        col_qty       = _find(["quantity"]) or _find(["qty"]) or _find(["পরিমাণ"])
        col_unit      = _find(["unit"], exclude=["per unit", "vat", "sd", "rate", "mrp"])
        col_mrp       = _find(["mrp"])
        col_rate      = (
            _find(["per unit rate"]) or _find(["per unit price"]) or
            _find(["unit rate"]) or _find(["unit price"], ["vat", "sd", "total"]) or
            _find(["rate"], ["vat", "sd", "total", "per unit", "mrp"]) or
            _find(["রেট"])
        )
        col_total_amt = _find(["total amount"], ["vat", "tax", "incl", "discount"])
        col_discount  = _find(["discount"]) or _find(["ডিসকাউন্ট"])
        col_sd_rate   = _find(["sd", "rate"])
        col_sd_amt    = _find(["sd", "amount"])
        col_vat_rate  = _find(["vat", "rate"]) or _find(["tax", "rate"])
        col_vat_amt   = _find(["vat", "amount"]) or _find(["tax", "amount"])
        col_line_total = _find(["total value"]) or _find(["total", "discount", "vat"]) or _find(["grand total"]) or _find(["মোট"], ["ডিসকাউন্ট"])

        if col_name is None:
            raise ValidationError(
                "Could not find a name/product column. "
                "Expected a header like 'Product / Service Name'."
            )
        if col_qty is None:
            raise ValidationError(
                "Could not find a quantity column. "
                "Expected a header like 'Quantity'."
            )

        # Build ingredient lookup: alias text → Ingredient, name → Ingredient
        all_ingredients = list(Ingredient.objects.filter(is_active=True))
        ing_by_name = {ing.name.lower(): ing for ing in all_ingredients}
        ing_by_alias = {}
        for alias in SupplierProductAlias.objects.filter(is_active=True).select_related("ingredient"):
            ing_by_alias[alias.alias_text.lower()] = alias.ingredient

        def _match_ingredient(text: str):
            lower = text.strip().lower()
            if lower in ing_by_alias:
                return ing_by_alias[lower]
            if lower in ing_by_name:
                return ing_by_name[lower]
            try:
                from rapidfuzz import fuzz, process as rfprocess
                choices = list(ing_by_name.keys())
                result = rfprocess.extractOne(lower, choices, scorer=fuzz.token_sort_ratio)
                if result and result[1] >= 80:
                    return ing_by_name[result[0]]
            except ImportError:
                pass
            return None

        def _decimal(val):
            if val is None:
                return None
            s = str(val).strip().replace(",", "")
            if not s or s in ("-", ""):
                return None
            try:
                d = Decimal(s)
                return d if d >= 0 else None
            except Exception:
                return None

        def _cell(row, col_idx):
            if col_idx is None or col_idx >= len(row):
                return None
            return row[col_idx].value  # row contains Cell objects

        def _pct_cell(row, col_idx):
            """Like _cell but multiplies by 100 when the cell is percentage-formatted.
            openpyxl stores 15% cells as 0.15; this restores the display value (15)."""
            if col_idx is None or col_idx >= len(row):
                return None
            cell = row[col_idx]
            val = cell.value
            if val is None:
                return None
            if isinstance(val, (int, float)) and cell.number_format and "%" in cell.number_format:
                return val * 100
            return val

        _SKIP = frozenset(["grand total", "sub total", "subtotal", "vat total", "tax total", "total"])

        record.items.filter(source=LineSource.SLIP_EXTRACTED).delete()

        created = 0
        unresolved = []

        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue

            name_raw = _cell(row, col_name)
            if name_raw is None or str(name_raw).strip() == "":
                continue
            name_text = str(name_raw).strip()
            if name_text.lower() in _SKIP or any(kw in name_text.lower() for kw in ["grand total", "sub total", "subtotal"]):
                continue

            qty = _decimal(_cell(row, col_qty)) or Decimal("0")

            unit_raw = _cell(row, col_unit)
            unit = UnitCaptured.PIECE if (
                unit_raw and str(unit_raw).strip().upper() == "PIECE"
            ) else UnitCaptured.PACK

            sku_code   = str(_cell(row, col_sku) or "").strip()
            mrp        = _decimal(_cell(row, col_mrp))
            rate       = _decimal(_cell(row, col_rate))
            total_amt  = _decimal(_cell(row, col_total_amt))
            discount   = _decimal(_cell(row, col_discount))
            sd_rate    = _decimal(_pct_cell(row, col_sd_rate))
            sd_amt     = _decimal(_cell(row, col_sd_amt))
            vat_rate   = _decimal(_pct_cell(row, col_vat_rate))
            vat_amt    = _decimal(_cell(row, col_vat_amt))
            line_total = _decimal(_cell(row, col_line_total))

            # Fallback: derive line_total from rate/discount when column absent.
            if line_total is None and rate and qty:
                line_total = (rate * qty - (discount or Decimal("0"))).quantize(Decimal("0.01"))

            unit_price = None
            if line_total and qty:
                try:
                    unit_price = (line_total / qty).quantize(Decimal("0.0001"))
                except Exception:
                    pass

            ingredient = _match_ingredient(name_text)
            pack_definition_id = None
            if ingredient:
                active = ingredient.active_pack()
                pack_definition_id = active.id if active else None

            StockInItem.objects.create(
                stock_in_record=record,
                ingredient=ingredient,
                raw_extracted_text=name_text,
                source=LineSource.SLIP_EXTRACTED,
                unit_captured=unit,
                extracted_quantity=qty,
                confirmed_quantity=qty,
                pack_definition_id=pack_definition_id,
                sku_code=sku_code,
                mrp=mrp,
                unit_price=unit_price,
                rate=rate,
                total_amount=total_amt,
                discount=discount,
                sd_rate=sd_rate,
                sd_amount=sd_amt,
                vat_rate=vat_rate,
                vat_amount=vat_amt,
                line_total=line_total,
            )
            created += 1
            if ingredient is None:
                unresolved.append(name_text)

        record = self.get_queryset().get(pk=record.pk)
        data = self.get_serializer(record).data
        data["imported_count"] = created
        data["unresolved_names"] = unresolved
        return Response(data)


class PreparationLogViewSet(viewsets.ModelViewSet):
    queryset = PreparationLog.objects.select_related("product", "outlet")
    serializer_class = PreparationLogSerializer

    def get_serializer_class(self):
        if self.action == "list" and self.request.query_params.get("slim") == "1":
            return PrepLogSlimSerializer
        return PreparationLogSerializer

    def get_queryset(self):
        if self.action == "list" and self.request.query_params.get("slim") == "1":
            qs = PreparationLog.objects.select_related("product")
        else:
            qs = super().get_queryset()
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        date_param = self.request.query_params.get("date")
        if date_param:
            from django.db.models import Q as DQ
            qs = qs.filter(DQ(op_date=date_param) | DQ(op_date__isnull=True, timestamp__date=date_param))
        elif self.request.query_params.get("today") == "true":
            local_today = timezone.localdate()
            from django.db.models import Q as DQ
            qs = qs.filter(DQ(op_date=local_today) | DQ(op_date__isnull=True, timestamp__date=local_today))
        return qs

    def _sync_closing_available(self, outlet, product, delta, op_date=None):
        """If an open closing exists for today with a stock count for this product,
        shift available_pieces by delta and recompute flags / walk-in lines."""
        from closing.models import DailyClosing, DailyClosingStockCount
        from closing.services import recompute_closing as _recompute_closing
        date = op_date or timezone.localdate()
        closing = (
            DailyClosing.objects
            .filter(outlet=outlet, closing_date=date)
            .exclude(status="LOCKED")
            .first()
        )
        if not closing:
            return
        updated = DailyClosingStockCount.objects.filter(
            daily_closing=closing, product=product
        ).update(available_pieces=models.F("available_pieces") + delta)
        if updated:
            _recompute_closing(closing)

    def perform_create(self, serializer):
        data = serializer.validated_data
        product = data["product"]
        outlet = data["outlet"]
        source = data.get("source", PrepSource.FRESH)
        op_date = data.get("op_date")

        if source == PrepSource.CARRIED_FORWARD:
            leftover = int(data.get("leftover_available_pieces") or 0)
            pieces = int(data.get("pieces_prepared") or leftover)
            pieces = min(pieces, leftover) if leftover else pieces
            wastage = max(0, leftover - pieces)
            instance = serializer.save(
                logged_by=self.request.user,
                pieces_prepared=pieces,
                wastage_pieces=wastage,
                leftover_available_pieces=leftover,
                op_date=op_date,
            )
            DisplayStock.adjust(outlet, product, pieces)
            return instance

        # FRESH
        unit = data.get("prep_unit", PrepUnit.PIECE)
        recipes = list(product.recipes.select_related("ingredient"))
        if unit == PrepUnit.PACK:
            # Find the primary ingredient: explicit is_primary flag, else the
            # only recipe ingredient (backward compat for single-ingredient products).
            primary = next((r for r in recipes if r.is_primary), None)
            if primary is None:
                if len(recipes) == 1:
                    primary = recipes[0]
                else:
                    raise ValidationError(
                        "PACK entry requires a designated primary ingredient. "
                        "Mark one ingredient as primary in the recipe editor."
                    )
            packs_used = Decimal(str(data["packs_used"]))
            pack = primary.ingredient.active_pack()
            if not pack:
                raise ValidationError("No active pack definition for the primary ingredient.")
            # pieces_per_pack is in the ingredient's base_unit; divide by
            # quantity_per_unit to get finished product pieces.
            primary_qty = primary.quantity_per_unit or Decimal("1")
            pieces = int(packs_used * pack.pieces_per_pack / primary_qty)
            instance = serializer.save(
                logged_by=self.request.user, pieces_prepared=pieces, op_date=op_date
            )
            # Deduct ALL recipe ingredients proportionally based on pieces prepared.
            for r in recipes:
                base_units = Decimal(pieces) * r.quantity_per_unit
                RawStock.adjust(outlet, r.ingredient, -base_units)
            # Deduct any prepared product components
            for comp_row in product.product_recipe_components.select_related("component_product"):
                qty = int(Decimal(pieces) * comp_row.quantity_per_unit)
                DisplayStock.adjust(outlet, comp_row.component_product, -qty)
        else:  # PIECE
            pieces = int(data["pieces_prepared"])
            instance = serializer.save(logged_by=self.request.user, pieces_prepared=pieces, op_date=op_date)
            consume_for_preparation(outlet, product, pieces)

        DisplayStock.adjust(outlet, product, pieces)
        self._sync_closing_available(outlet, product, pieces, op_date)
        return instance

    def perform_destroy(self, instance):
        outlet = instance.outlet
        product = instance.product
        pieces = instance.pieces_prepared

        # Block deletion when a downstream product was prepared using this
        # product as a display-stock component on the same day.
        component_links = list(
            product.used_as_prep_component.select_related("product").all()
        )
        if component_links:
            from django.db.models import Q as DQ
            op_date = instance.op_date or instance.timestamp.date()
            downstream_products = [link.product for link in component_links]
            blocking_logs = PreparationLog.objects.filter(
                outlet=outlet,
                product__in=downstream_products,
            ).filter(
                DQ(op_date=op_date)
                | DQ(op_date__isnull=True, timestamp__date=op_date)
            ).select_related("product")
            if blocking_logs.exists():
                names = ", ".join(
                    sorted({log.product.name for log in blocking_logs})
                )
                raise ValidationError(
                    f"Cannot delete this prep log — {names} "
                    f"{'has' if names.count(',') == 0 else 'have'} already been "
                    f"prepared using {product.name} pieces today. "
                    f"Delete {'that' if names.count(',') == 0 else 'those'} prep log(s) first."
                )

        op_date = instance.op_date or instance.timestamp.date()
        if instance.source == PrepSource.FRESH:
            restock_from_preparation(outlet, product, pieces)
        # For CARRIED_FORWARD, no raw stock was touched — only undo DisplayStock.
        DisplayStock.adjust(outlet, product, -pieces)
        instance.delete()
        self._sync_closing_available(outlet, product, -pieces, op_date)


class RawStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = RawStock.objects.select_related("ingredient", "outlet").prefetch_related(
        "ingredient__aliases",
        "ingredient__pack_definitions",
    )
    serializer_class = RawStockSerializer

    def get_serializer_class(self):
        if self.request.query_params.get("slim") == "1":
            return RawStockPrepSerializer
        return RawStockSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        if self.get_serializer_class() is RawStockSerializer:
            from catalog.utils import build_ingredient_category_map, build_ingredient_product_map
            ctx["ingredient_category_map"] = build_ingredient_category_map()
            ctx["ingredient_product_map"] = build_ingredient_product_map()
        return ctx

    def get_queryset(self):
        if self.request.query_params.get("slim") == "1":
            # Skip the aliases prefetch — RawStockPrepSerializer doesn't use it.
            qs = RawStock.objects.select_related("ingredient").prefetch_related(
                "ingredient__pack_definitions"
            )
        else:
            qs = super().get_queryset()
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        return qs


class DisplayStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DisplayStock.objects.select_related("product", "outlet")
    serializer_class = DisplayStockSerializer

    def get_serializer_class(self):
        if self.request.query_params.get("slim") == "1":
            return DisplayStockPrepSerializer
        return DisplayStockSerializer

    def get_queryset(self):
        # Slim mode: drop select_related("product") — DisplayStockPrepSerializer doesn't access it.
        qs = DisplayStock.objects.all() if self.request.query_params.get("slim") == "1" else super().get_queryset()
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        return qs


def _initialize_direct_stock(outlet):
    """Set DisplayStock for non-prep (beverage) products from current RawStock.

    Called when the operating day moves to IN_PROGRESS and on every stock-in
    approval, so DisplayStock stays in sync with whatever is physically available.
    """
    from catalog.models import Product
    for product in Product.objects.filter(
        requires_preparation=False, is_active=True
    ).prefetch_related("recipes__ingredient"):
        recipes = list(product.recipes.all())
        if not recipes:
            continue
        cap = None
        for r in recipes:
            raw = RawStock.objects.filter(outlet=outlet, ingredient=r.ingredient).first()
            available = raw.quantity_available if raw else Decimal("0")
            qty_per = r.quantity_per_unit or Decimal("1")
            pieces = int(available / qty_per)
            cap = pieces if cap is None else min(cap, pieces)
        ds, _ = DisplayStock.objects.get_or_create(outlet=outlet, product=product)
        ds.pieces_available = cap or 0
        ds.save(update_fields=["pieces_available"])


class OperatingDayViewSet(viewsets.ReadOnlyModelViewSet):
    """Gates the staff daily flow. Read + custom transitions (start, confirm)."""

    queryset = OperatingDay.objects.select_related("outlet", "started_by").prefetch_related(
        "stock_checks__ingredient__aliases",
    )
    serializer_class = OperatingDaySerializer

    def get_queryset(self):
        qs = super().get_queryset()
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        date = self.request.query_params.get("date")
        if date:
            qs = qs.filter(date=date)
        return qs

    def _outlet(self, request):
        outlet_id = request.data.get("outlet") or request.query_params.get("outlet")
        if outlet_id:
            from catalog.models import Outlet

            return Outlet.objects.get(pk=outlet_id)
        return request.user.outlet

    @action(detail=False, methods=["get", "post"])
    def today(self, request):
        """Get (or lazily create) an OperatingDay. Accepts ?date=YYYY-MM-DD;
        defaults to today when omitted. If a previous day is not yet CLOSED,
        returns it instead of creating a new day (day boundary enforcement).
        Pass ?slim=1 to skip stock_checks (layout nav calls only need status fields)."""
        from datetime import date as dt_date
        outlet = self._outlet(request)
        on_date = None
        raw_date = request.query_params.get("date")
        if raw_date:
            try:
                on_date = dt_date.fromisoformat(raw_date)
            except ValueError:
                pass

        slim = request.query_params.get("slim") == "1"
        serializer_class = OperatingDaySlimSerializer if slim else OperatingDaySerializer

        # Day boundary guard: applies when requesting today (explicit or implicit).
        # Explicit past dates (backdating for historical entry) bypass the guard.
        actual_date = timezone.localdate()
        if on_date is None or on_date == actual_date:
            latest = OperatingDay.objects.filter(outlet=outlet).order_by("-date").first()
            if latest and latest.date < actual_date and latest.status != OperatingDayStatus.CLOSED:
                return Response(serializer_class(latest, context={"request": request}).data)

        day = get_or_create_today(outlet, on_date=on_date)
        return Response(serializer_class(day, context={"request": request}).data)

    @action(detail=True, methods=["post"], url_path="force-close")
    def force_close(self, request, pk=None):
        """Force-close a stuck operating day that was never properly closed
        (e.g., outlet didn't operate that day). Moves status directly to CLOSED."""
        day = self.get_object()
        if day.status == OperatingDayStatus.CLOSED:
            return Response(self.get_serializer(day).data)
        day.status = OperatingDayStatus.CLOSED
        day.save(update_fields=["status"])
        return Response(self.get_serializer(day).data)

    @action(detail=True, methods=["post"])
    def start(self, request, pk=None):
        day = self.get_object()
        if day.status == OperatingDayStatus.NOT_STARTED:
            day.status = OperatingDayStatus.NOT_STARTED  # stays; stock check advances it
            day.started_by = request.user
            day.started_at = timezone.now()
            day.save()
        return Response(self.get_serializer(day).data)

    @action(detail=True, methods=["get"], url_path="day-start-stock")
    def day_start_stock(self, request, pk=None):
        """Rows to reconcile: every RECIPE_LINKED ingredient with its carried
        RawStock balance, merged with any already-saved confirmations."""
        from catalog.utils import build_ingredient_category_map, build_ingredient_product_map, resolve_ingredient_group
        day = self.get_object()
        existing = {c.ingredient_id: c for c in day.stock_checks.select_related("ingredient")}
        category_map = build_ingredient_category_map()
        product_map = build_ingredient_product_map()
        rows = []
        ingredients = Ingredient.objects.filter(
            is_active=True, tracking_mode=TrackingMode.RECIPE_LINKED
        )
        for ing in ingredients:
            rs = RawStock.objects.filter(outlet=day.outlet, ingredient=ing).first()
            carried = rs.quantity_available if rs else Decimal("0")
            check = existing.get(ing.id)
            active_pack = ing.active_pack()
            alias = ing.aliases.filter(is_active=True).first()
            rows.append(
                {
                    "ingredient": ing.id,
                    "ingredient_name": ing.name,
                    "ingredient_display_name": alias.alias_text if alias else ing.name,
                    "ingredient_group": resolve_ingredient_group(ing, category_map),
                    "primary_product": product_map.get(ing.id, ""),
                    "base_unit": ing.base_unit,
                    "pieces_per_pack": str(active_pack.pieces_per_pack) if active_pack else None,
                    "system_carried_qty": carried,
                    "confirmed_qty": check.confirmed_qty if check else carried,
                    "discrepancy_reason": check.discrepancy_reason if check else "",
                    "note": check.note if check else "",
                }
            )
        return Response(rows)

    @action(detail=True, methods=["post"], url_path="confirm-stock")
    def confirm_stock(self, request, pk=None):
        """Save DayStartStockCheck rows, set RawStock to confirmed values, advance
        to STOCK_CONFIRMED. Any discrepancy needs a reason. Body: {items:[...]}."""
        day = self.get_object()
        for row in request.data.get("items", []):
            ing = Ingredient.objects.get(pk=row["ingredient"])
            system = Decimal(str(row["system_carried_qty"]))
            confirmed = Decimal(str(row["confirmed_qty"]))
            reason = row.get("discrepancy_reason", "")
            if system != confirmed and not reason:
                raise ValidationError(
                    f"'{ing.name}' has a discrepancy — a reason is required."
                )
            obj, _ = DayStartStockCheck.objects.get_or_create(
                operating_day=day, ingredient=ing
            )
            obj.system_carried_qty = system
            obj.confirmed_qty = confirmed
            obj.discrepancy_reason = reason
            obj.note = row.get("note", "")
            obj.save()
            RawStock.set_to(day.outlet, ing, confirmed)
        if day.status == OperatingDayStatus.NOT_STARTED:
            # Reset DisplayStock to 0 for all products so today starts clean.
            # Carry-forward will add back yesterday's genuine leftovers.
            DisplayStock.objects.filter(outlet=day.outlet).update(pieces_available=0)
            day.status = OperatingDayStatus.STOCK_CONFIRMED
        day.stock_confirmed_at = timezone.now()
        if not day.started_at:
            day.started_by = request.user
            day.started_at = timezone.now()
        day.save()
        return Response(self.get_serializer(day).data)

    @action(detail=True, methods=["get"], url_path="carry-forward")
    def carry_forward_list(self, request, pk=None):
        """Yesterday's leftovers to move into today's display stock."""
        day = self.get_object()
        rows = []
        for count in carry_forward_candidates(day):
            rows.append(
                {
                    "stock_count": count.id,
                    "product": count.product_id,
                    "product_name": count.product.name,
                    "leftover_available_pieces": count.remains_pieces,
                    "pieces_prepared": count.remains_pieces,
                }
            )
        return Response(rows)

    @action(detail=True, methods=["post"], url_path="confirm-carry-forward")
    def confirm_carry_forward(self, request, pk=None):
        """Create CARRIED_FORWARD PreparationLog rows (no RawStock change) and
        advance to IN_PROGRESS. Partial moves auto-fill wastage. Body:{items:[...]}."""
        from closing.models import DailyClosingStockCount

        day = self.get_object()
        if day.status == OperatingDayStatus.NOT_STARTED:
            raise ValidationError("Confirm day-start stock first.")

        # Idempotent: undo any previous carry-forward for this day.
        existing = PreparationLog.objects.filter(
            outlet=day.outlet,
            source=PrepSource.CARRIED_FORWARD,
            op_date=day.date,
        )
        for prev in existing:
            if prev.pieces_prepared:
                DisplayStock.adjust(day.outlet, prev.product, -prev.pieces_prepared)
        existing.delete()

        # Do NOT reset DisplayStock here — the undo loop above already reversed the
        # previous carry-forward contribution precisely. A blanket reset would wipe
        # fresh prep contributions logged between the first and second carry-forward
        # submissions, leaving DisplayStock lower than the sum of today's prep logs.

        for row in request.data.get("items", []):
            count = DailyClosingStockCount.objects.get(pk=row["stock_count"])
            leftover = int(row.get("leftover_available_pieces", count.remains_pieces))
            pieces = int(row.get("pieces_prepared", leftover))
            pieces = max(0, min(pieces, leftover))
            wastage = leftover - pieces
            PreparationLog.objects.create(
                outlet=day.outlet,
                logged_by=request.user,
                product=count.product,
                source=PrepSource.CARRIED_FORWARD,
                carried_forward_from=count,
                leftover_available_pieces=leftover,
                pieces_prepared=pieces,
                wastage_pieces=wastage,
                op_date=day.date,
            )
            if pieces:
                DisplayStock.adjust(day.outlet, count.product, pieces)
        day.status = OperatingDayStatus.IN_PROGRESS
        day.carry_forward_confirmed_at = timezone.now()
        day.save()
        _initialize_direct_stock(day.outlet)
        return Response(self.get_serializer(day).data)


class PeriodicStockCheckViewSet(viewsets.ModelViewSet):
    """Packaging & supplies — staff reports what's left; consumption is inferred."""

    queryset = PeriodicStockCheck.objects.select_related("ingredient", "outlet", "checked_by")
    serializer_class = PeriodicStockCheckSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        ingredient = self.request.query_params.get("ingredient")
        if ingredient:
            qs = qs.filter(ingredient_id=ingredient)
        if self.request.query_params.get("latest_per_ingredient") == "true":
            # Keep only the newest row per ingredient (small dataset — do it in py).
            seen, keep = set(), []
            for row in qs:
                if row.ingredient_id in seen:
                    continue
                seen.add(row.ingredient_id)
                keep.append(row.id)
            qs = qs.filter(id__in=keep)
        return qs

    def _record(self, outlet, ingredient, counted_qty, note, user):
        prev = (
            PeriodicStockCheck.objects.filter(outlet=outlet, ingredient=ingredient)
            .order_by("-checked_at")
            .first()
        )
        since = prev.checked_at if prev else None
        stock_in = stock_in_since(outlet, ingredient, since)
        prev_qty = prev.counted_qty if prev else Decimal("0")
        consumed = prev_qty + stock_in - counted_qty
        return PeriodicStockCheck.objects.create(
            outlet=outlet,
            ingredient=ingredient,
            checked_by=user,
            counted_qty=counted_qty,
            stock_in_since_last_check=stock_in,
            consumed_since_last_check=consumed,
            note=note,
        )

    def create(self, request, *args, **kwargs):
        """Full recount. Body: {outlet, ingredient, counted_qty, note?}."""
        ingredient = Ingredient.objects.get(pk=request.data["ingredient"])
        outlet = _outlet_obj(request.data.get("outlet") or request.user.outlet_id)
        obj = self._record(
            outlet,
            ingredient,
            Decimal(str(request.data["counted_qty"])),
            request.data.get("note", ""),
            request.user,
        )
        return Response(self.get_serializer(obj).data, status=201)

    @action(detail=False, methods=["post"], url_path="bundle-finished")
    def bundle_finished(self, request, pk=None):
        """One-tap 'used the last of a bundle' — subtracts the pack size from the
        last count. Body: {outlet, ingredient}."""
        ingredient = Ingredient.objects.get(pk=request.data["ingredient"])
        outlet = _outlet_obj(request.data.get("outlet") or request.user.outlet_id)
        prev = (
            PeriodicStockCheck.objects.filter(outlet=outlet, ingredient=ingredient)
            .order_by("-checked_at")
            .first()
        )
        pack = ingredient.active_pack()
        bundle = pack.pieces_per_pack if pack else Decimal("0")
        prev_qty = prev.counted_qty if prev else bundle
        counted = max(Decimal("0"), prev_qty - bundle)
        obj = self._record(
            outlet, ingredient, counted, "Bundle finished (−1 pack)", request.user
        )
        return Response(self.get_serializer(obj).data, status=201)

    @action(detail=False, methods=["get"], url_path="levels")
    def levels(self, request):
        """Effective current qty for every PERIODIC_COUNT ingredient.
        Returns the latest counted_qty when a check exists, otherwise derives
        the quantity from all approved stock-in records (bootstrap case)."""
        outlet = _outlet_obj(
            request.query_params.get("outlet") or request.user.outlet_id
        )
        ingredients = Ingredient.objects.filter(
            tracking_mode=TrackingMode.PERIODIC_COUNT, is_active=True
        ).prefetch_related("aliases")
        # Latest check per ingredient (Python dedup — SQLite-safe, no DISTINCT ON).
        checks_qs = (
            PeriodicStockCheck.objects
            .filter(outlet=outlet)
            .order_by("-checked_at")
            .select_related("ingredient")
        )
        latest_by_ing = {}
        for c in checks_qs:
            if c.ingredient_id not in latest_by_ing:
                latest_by_ing[c.ingredient_id] = c

        from catalog.utils import build_ingredient_category_map, resolve_ingredient_group
        category_map = build_ingredient_category_map()
        result = []
        for ing in ingredients:
            check = latest_by_ing.get(ing.id)
            pack = ing.active_pack()
            if check:
                # Add any stock-in received since the last physical count.
                new_stock = stock_in_since(outlet, ing, since_dt=check.checked_at)
                current_qty = str(check.counted_qty + new_stock)
                source = "counted" if new_stock == 0 else "counted_plus_stock_in"
                last_at = check.checked_at.isoformat()
            else:
                total = stock_in_since(outlet, ing, since_dt=None)
                current_qty = str(total)
                source = "stock_in_derived"
                last_at = None
            alias = ing.aliases.filter(is_active=True).first()
            cost_per_base_unit = ing.cost_per_base_unit
            result.append({
                "ingredient": ing.id,
                "ingredient_name": ing.name,
                "ingredient_display_name": alias.alias_text if alias else ing.name,
                "ingredient_group": resolve_ingredient_group(ing, category_map),
                "base_unit": ing.base_unit,
                "pieces_per_pack": str(pack.pieces_per_pack) if pack else None,
                "current_qty": current_qty,
                "cost_per_base_unit": str(cost_per_base_unit) if cost_per_base_unit is not None else None,
                "source": source,
                "last_checked_at": last_at,
            })
        return Response(result)


def _outlet_obj(outlet_or_id):
    from catalog.models import Outlet

    if outlet_or_id is None:
        return None
    if hasattr(outlet_or_id, "pk"):
        return outlet_or_id
    return Outlet.objects.get(pk=outlet_or_id)


from rest_framework.views import APIView as _HomeSummaryBase
from rest_framework.permissions import IsAuthenticated as _HomeSummaryAuth


class StaffHomeSummaryView(_HomeSummaryBase):
    """Single endpoint aggregating all data the staff home page needs.

    Replaces four separate API calls (/outlets/, /display-stock/, /stock-in/,
    /daily-closings/) with one server-side computed response.
    """

    permission_classes = [_HomeSummaryAuth]

    def get(self, request):
        import datetime
        from catalog.models import Outlet
        from closing.models import DailyClosing

        outlet_id = int(request.query_params.get("outlet", 1))
        date_str = request.query_params.get("date", str(datetime.date.today()))

        # Outlet — only the one flag needed on the home page
        try:
            outlet_obj = Outlet.objects.only("allow_staff_date_selection").get(id=outlet_id)
            allow_date_selection = outlet_obj.allow_staff_date_selection
        except Outlet.DoesNotExist:
            allow_date_selection = True

        # Raw stock — all ingredient balances for the outlet.
        # active_pack() is prefetch-cache-aware, so this stays 2 queries total.
        raw_stocks = list(
            RawStock.objects
            .filter(outlet_id=outlet_id)
            .select_related("ingredient")
            .prefetch_related("ingredient__pack_definitions")
        )
        raw_stock_value = sum(
            s.quantity_available * s.ingredient.cost_per_base_unit
            for s in raw_stocks
            if s.quantity_available > 0
        )
        raw_category_count = len({
            s.ingredient.group
            for s in raw_stocks
            if s.quantity_available > 0 and s.ingredient.group
        })
        raw_stock_out_count = sum(1 for s in raw_stocks if s.quantity_available <= 0)

        # Stock-in — one annotated query covers draft detection and latest status
        si_records = list(
            StockInRecord.objects
            .filter(outlet_id=outlet_id, stock_in_date=date_str)
            .annotate(item_count=Count("items", distinct=True))
            .order_by("-created_at")
        )
        draft_si = next((r for r in si_records if r.status == StockInStatus.DRAFT), None)
        latest_si = si_records[0] if si_records else None
        stock_in_data = None
        if latest_si:
            stock_in_data = {
                "draft_item_count": draft_si.item_count if draft_si else None,
                "latest_status": latest_si.status,
            }

        # Closing — prefetch what total_sale / has_flag need
        closing = (
            DailyClosing.objects
            .prefetch_related("sales_lines", "channel_discounts", "payments")
            .filter(outlet_id=outlet_id, closing_date=date_str)
            .first()
        )
        closing_data = None
        if closing:
            closing_data = {
                "id": closing.id,
                "status": closing.status,
                "total_sale": str(closing.total_sale),
                "has_flag": closing.has_flag,
            }

        return Response({
            "allow_staff_date_selection": allow_date_selection,
            "display_stock": {
                "raw_stock_value": str(raw_stock_value),
                "raw_category_count": raw_category_count,
                "raw_stock_out_count": raw_stock_out_count,
            },
            "stock_in": stock_in_data,
            "closing": closing_data,
        })


class FryerOilChangeViewSet(viewsets.ModelViewSet):
    serializer_class = FryerOilChangeSerializer
    http_method_names = ["get", "post", "head", "options"]

    def get_queryset(self):
        qs = FryerOilChange.objects.select_related("logged_by")
        outlet = self.request.query_params.get("outlet")
        if outlet:
            qs = qs.filter(outlet_id=outlet)
        pan = self.request.query_params.get("pan")
        if pan:
            qs = qs.filter(pan_number=pan)
        return qs

    def perform_create(self, serializer):
        serializer.save(logged_by=self.request.user)
